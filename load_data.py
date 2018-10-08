from openpyxl import load_workbook
import pandas as pd


def load_data_from_xlsx(s):
    wb = load_workbook(s)
    ws = wb.active
    data = []
    i = 1
    while ws.cell(row=i, column=1).value is not None:
        date = ws.cell(row=i, column=1).value
        H2 = ws.cell(row=i, column=2).value
        CH4 = ws.cell(row=i, column=3).value
        CO = ws.cell(row=i, column=4).value
        CO2 = ws.cell(row=i, column=5).value
        C2H6 = ws.cell(row=i, column=6).value
        C2H4 = ws.cell(row=i, column=7).value
        C2H2 = ws.cell(row=i, column=8).value
        data.append((date, H2, CH4, CO, CO2, C2H6, C2H4, C2H2))
        i += 1
    wb.close()
    d = pd.DataFrame(data, columns=('date', 'H2', 'CH4', 'CO', 'CO2', 'C2H6', 'C2H4', 'C2H2'))
    return d
